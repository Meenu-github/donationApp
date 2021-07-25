import streamlit as st
import openpyxl as pxl
import pandas as pd

def dataRetrieving(y):
    
    def type(selectRole):
        if selectRole == 'Hospital':
            exp = pd.read_excel('../donationApp/BloodDonation.xlsx')
            new = pxl.load_workbook('../donationApp/BloodDonation.xlsx')
            newsheet = new.active
            #sheet = exp.active
            #maxrow = sheet.max_row+1
            
            st.table(exp)
        
        if selectRole == 'Food Distributor':
            exp1 = pd.read_excel('../donationApp/FoodDonation.xlsx')
            new = pxl.load_workbook('../donationApp/FoodDonation.xlsx')
            newsheet = new.active
            #sheet = exp1.active
            #maxrow = sheet.max_row+1
            st.table(exp1)
        
        if selectRole == 'Orphanage':
            exp2 = pd.read_excel('../donationApp/BookDonation.xlsx')
            new = pxl.load_workbook('../donationApp/BookDonation.xlsx')
            newsheet = new.active
            #sheet = exp2.active
            #maxrow = sheet.max_row+1
            st.table(exp2)

    data = ['Yes, I want Donor data', 'No, I don\'t want Donor data']
    b = st.selectbox('Do you want data of blood donor :', data)
    if b == 'Yes, I want Blood Donor data':
        type(y)
    if b == 'No, I want Blood Donor data':
        st.info('YOU SELECTED NO.')