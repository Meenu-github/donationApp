import streamlit as st
import openpyxl as pxl
import pandas as pd

#import gspread
#from gspread_dataframe import get_as_dataframe 
#proj = pd.read_excel('proj.xlsx')
#exp = pxl.load_workbook('../donationApp/proj.xlsx')

projDemo = pd.read_excel('proj.xlsx')
exp = pxl.load_workbook('proj.xlsx')
sheet = exp.active
maxrow = sheet.max_row+1
#sheet = exp['Sheet2']

def type(selectRole):
    if selectRole == 'Hospital':
        exp = pd.read_excel('BloodDonation.xlsx')
        
        st.table(exp)
        

    if selectRole == 'Food Distributor':
        exp1 = pd.read_excel('FoodDonation.xlsx')
        
        st.table(exp1)
        

    if selectRole == 'Orphanage':
        exp2 = pd.read_excel('BookDonation.xlsx')
        
        st.table(exp2)
        


def loginPages():
    page_name = ['Register', 'Login']
    page = st.radio('Choose Register/Login', page_name)
    if page == 'Register':
        with st.form(key="Sign up"):
            userName = st.text_input("Username")
            email = st.text_input("E-mail Id")
            password = st.text_input("Password", type="password")
            conf_password = st.text_input('Confirm your password', type="password")
            role = ['Hospital', 'Food Distributor', 'Orphanage']
            selectRole = st.selectbox("Role", role)
            if password == conf_password:
                pass
            else:
                st.warning("password do not match with confirm password")
               

            sheet.cell(row=maxrow, column=1).value = userName
            sheet.cell(row=maxrow, column=2).value = email
            sheet.cell(row=maxrow, column=3).value = password
            sheet.cell(row=maxrow, column=4).value = selectRole

            submissionButton = st.form_submit_button(label="Sign up")
            if submissionButton == True:
                exp.save('proj.xlsx')
                st.success("Successfully sign up")
                
                st.info('Giving data of donor')

                type(selectRole)
                
                

    if page == 'Login' :
        with st.form(key="Login"):

            userName = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submissionButton = st.form_submit_button(label="Login")
            if submissionButton == True:

                for i in range(2, sheet.max_row):
                    if((sheet.cell(row=i, column=2).value == userName)):
                        if((sheet.cell(row=i, column=3).value == password)):
                            y = sheet.cell(row=i, column=4).value
                            st.success('Successfully Login')
                            
                            st.info('Giving data of donor')
                            type(y)
                            
                            
                            

                        else:
                            st.error(
                                "either username or password is incorrect")





