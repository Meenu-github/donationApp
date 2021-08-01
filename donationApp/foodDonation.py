import streamlit as st
import openpyxl as pxl
from PIL import Image
#import streamlit as st
import base64
import sys
import pyodbc as odbc
records = []
DRIVER = "SQL Server"
SERVER_NAME = "MEENU\SQLEXPRESS"
DATABASE_NAME="StreamLit"
cnxn = f"""
    Driver={{{DRIVER}}};
    Server={SERVER_NAME};
    Database={DATABASE_NAME};
    Trusted_Connection=yes;
"""

#wb = pxl.load_workbook('FoodDonation.xlsx')
#ws = wb.active
#maxrow= ws.max_row+1
#ws.title = "fooddonation"


def foodDonate() :
    st.title("Food Donation")
    main_bg = "food.gif"
    main_bg_ext = "gif"
    st.markdown(
    f"""
    <style>
    .reportview-container {{
        background: url(data:image/{main_bg_ext};base64,{base64.b64encode(open(main_bg, "rb").read()).decode()})
        
    }}
    </style>
    """,
    unsafe_allow_html=True
)
    st.markdown("FOOD DONATION")
    img = Image.open("FoodDonation.jpg")
    st.image(img, caption='Food Donation',width=500)
    st.title("Welcome to the food donation page, your donated food can bring hope in someones life of survival.\nCome let us donate food for needy one.\nYou don't have to walk and donate it you just have to register yourself and we will pick the food from your house address that will be provided.")
    st.write("Here if you are willing to donate food\n you have to register yourself.")
    with st.form(key="Register for food donation"):
        namefood = st.text_input("Enter your name : ")
        foodaddress = st.text_input("Enter your address please : ")
        food_phone = st.text_input("Enter your phone  Number : ")
        
        foodsubmission = st.form_submit_button(label="Submit")
        if foodsubmission==True:
            #ws.cell(row=maxrow,column=1).value = namefood
            #ws.cell(row=maxrow,column=2).value = foodaddress
            #ws.cell(row=maxrow,column=3).value = food_phone
            
            #wb.save('FoodDonation.xlsx')
            records.append([namefood,foodaddress,food_phone])
            addData()
            st.success("Successfully registered for food donation")
        else:
            st.info("Please submit the form.")

def addData():
    try:
        conn = odbc.connect(cnxn)
    except Exception as e:
        print(e)
        print("task is terminated")
        sys.exit()
    else:
        cursor = conn.cursor()
    insert_statement = """
        INSERT INTO Food_Donation
        VALUES (?, ?, ?)
    """
    try:
        for record in records:
            print(record)
            cursor.execute(insert_statement, record)
    except Exception as e:
        cursor.rollback()
        print(e.value)

    else:
        print("Successfully inserted")
        cursor.commit()
        cursor.close()

    finally:
        if conn.connected ==1:
            print("connection closed")
            conn.closed()




        

